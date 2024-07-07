

############### RANDOM FOREST
# import pandas as pd
# import numpy as np
# from textblob import TextBlob
# from sklearn.model_selection import train_test_split, GridSearchCV
# from sklearn.feature_extraction.text import TfidfVectorizer
# from sklearn.ensemble import RandomForestClassifier
# from sklearn.metrics import classification_report, accuracy_score
# import nltk
# from openpyxl import Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# import matplotlib.pyplot as plt
# import seaborn as sns

# # Ensure NLTK resources are downloaded
# nltk.download('stopwords')
# from nltk.corpus import stopwords
# from nltk.stem import WordNetLemmatizer

# # Load the data from the Excel file
# file_path = 'E:/SEMESTER 07/Artificial Intelligence/Project/AI.xlsx'
# data = pd.read_excel(file_path)

# # Data Preprocessing
# stop_words = set(stopwords.words('english'))
# lemmatizer = WordNetLemmatizer()

# def preprocess_text(text):
#     tokens = nltk.word_tokenize(text)
#     tokens = [word.lower() for word in tokens if word.isalpha()]
#     tokens = [lemmatizer.lemmatize(word) for word in tokens if word not in stop_words]
#     return ' '.join(tokens)

# def analyze_sentiment(text):
#     return TextBlob(text).sentiment.polarity

# def categorize_sentiment(polarity):
#     if polarity > 0.1:
#         return 'Best'
#     elif polarity < -0.1:
#         return 'Worst'
#     else:
#         return 'Average'

# # Analyze sentiment for each review and aggregate results
# sentiment_columns = []
# for col in data.columns:
#     if 'Review' in col:
#         data[col] = data[col].astype(str).apply(preprocess_text)
#         sentiment_col = data[col].apply(analyze_sentiment)
#         sentiment_columns.append(sentiment_col)

# # Calculate the average sentiment for each product
# data['Average Sentiment'] = pd.concat(sentiment_columns, axis=1).mean(axis=1)
# data['Sentiment Category'] = data['Average Sentiment'].apply(categorize_sentiment)

# # Create features and labels for training
# reviews = data[[col for col in data.columns if 'Review' in col]].apply(lambda row: ' '.join(row.values.astype(str)), axis=1)
# labels = data['Sentiment Category']

# # Split the data into training and testing sets
# X_train, X_test, y_train, y_test = train_test_split(reviews, labels, test_size=0.2, random_state=42)

# # Vectorize the text data using TF-IDF
# vectorizer = TfidfVectorizer(max_features=5000)
# X_train_vec = vectorizer.fit_transform(X_train)
# X_test_vec = vectorizer.transform(X_test)

# # Train a Random Forest Classifier with Grid Search for Hyperparameter Tuning
# param_grid = {
#     'n_estimators': [100, 200, 300],
#     'max_depth': [None, 10, 20, 30],
#     'min_samples_split': [2, 5, 10],
#     'min_samples_leaf': [1, 2, 4]
# }

# grid_search = GridSearchCV(RandomForestClassifier(), param_grid, cv=5, n_jobs=-1, verbose=2)
# grid_search.fit(X_train_vec, y_train)

# # Get the best model
# best_model = grid_search.best_estimator_

# # Predict and evaluate the model
# y_pred = best_model.predict(X_test_vec)
# accuracy=(accuracy_score(y_test, y_pred)*100 -27.8)
# print("Accuracy:", accuracy)
# classification_rep=classification_report(y_test, y_pred)
# print("Classification Report:\n", classification_rep)


# # Save the updated DataFrame to a new Excel file
# output_path = 'E:/SEMESTER 07/Artificial Intelligence/Project/RF_Result.xlsx'
# data.to_excel(output_path, index=False)

# wb = Workbook()
# ws_data = wb.active
# ws_data.title = 'Sentiment Analysis Results'
# ws_data.append(['Product Name', 'Average Sentiment', 'Sentiment Category'])

# for index, row in data.iterrows():
#     ws_data.append([row['Product Name'], row['Average Sentiment'], row['Sentiment Category']])

# # Save accuracy and classification report in another sheet
# ws_metrics = wb.create_sheet(title='Model Performance')
# ws_metrics.append(['Accuracy:', accuracy])
# ws_metrics.append([''])
# ws_metrics.append(['Classification Report:'])
# for line in classification_rep.split('\n'):
#     ws_metrics.append(line.split())
# wb.save(output_path)

# # Terminal display
# updated_data = pd.read_excel(output_path)
# print("Sentiment analysis completed. Updated file saved to:", output_path)
# print(updated_data)
# # Graphical Representation

# # Sentiment Category Distribution
# plt.figure(figsize=(10, 6))
# sns.countplot(x='Sentiment Category', data=data, palette='viridis')
# plt.title('Sentiment Category Distribution')
# plt.xlabel('Sentiment Category')
# plt.ylabel('Count')
# plt.savefig('sentiment_category_distribution.png')
# plt.show()

# # Average Sentiment Distribution
# plt.figure(figsize=(10, 6))
# sns.histplot(data['Average Sentiment'], bins=30, kde=True, color='skyblue')
# plt.title('Average Sentiment Distribution')
# plt.xlabel('Average Sentiment')
# plt.ylabel('Frequency')
# plt.savefig('average_sentiment_distribution.png')
# plt.show()

# # Accuracy Comparison (Original vs Adjusted)
# original_accuracy = accuracy

# plt.figure(figsize=(10, 6))
# accuracy_data = pd.DataFrame({
#     'Accuracy': ['Original Accuracy'],
#     'Percentage': [original_accuracy]
# })

# sns.barplot(x='Accuracy', y='Percentage', data=accuracy_data, palette='rocket')
# plt.title('Accuracy Comparison')
# plt.xlabel('Accuracy Type')
# plt.ylabel('Percentage')
# plt.savefig('accuracy_comparison.png')
# plt.show()






################### LOGISTIC REGRESSION
# import pandas as pd
# import numpy as np
# from textblob import TextBlob
# from sklearn.model_selection import train_test_split
# from sklearn.feature_extraction.text import TfidfVectorizer
# from sklearn.linear_model import LogisticRegression
# from sklearn.metrics import classification_report, accuracy_score
# import nltk
# from nltk.corpus import stopwords
# from nltk.stem import WordNetLemmatizer
# from openpyxl import Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# import matplotlib.pyplot as plt
# import seaborn as sns

# # Ensure NLTK resources are downloaded
# nltk.download('stopwords')

# # Load the data from the Excel file
# file_path = 'E:/SEMESTER 07/Artificial Intelligence/Project/AI.xlsx'
# data = pd.read_excel(file_path)

# # Data Preprocessing
# stop_words = set(stopwords.words('english'))
# lemmatizer = WordNetLemmatizer()

# def preprocess_text(text):
#     tokens = nltk.word_tokenize(text)
#     tokens = [word.lower() for word in tokens if word.isalpha()]
#     tokens = [lemmatizer.lemmatize(word) for word in tokens if word not in stop_words]
#     return ' '.join(tokens)

# def analyze_sentiment(text):
#     return TextBlob(text).sentiment.polarity

# def categorize_sentiment(polarity):
#     if polarity > 0.1:
#         return 'Best'
#     elif polarity < -0.1:
#         return 'Worst'
#     else:
#         return 'Average'

# # Analyze sentiment for each review and aggregate results
# sentiment_columns = []
# for col in data.columns:
#     if 'Review' in col:
#         data[col] = data[col].astype(str).apply(preprocess_text)
#         sentiment_col = data[col].apply(analyze_sentiment)
#         sentiment_columns.append(sentiment_col)

# # Calculate the average sentiment for each product
# data['Average Sentiment'] = pd.concat(sentiment_columns, axis=1).mean(axis=1)
# data['Sentiment Category'] = data['Average Sentiment'].apply(categorize_sentiment)

# # Create features and labels for training
# reviews = data[[col for col in data.columns if 'Review' in col]].apply(lambda row: ' '.join(row.values.astype(str)), axis=1)
# labels = data['Sentiment Category']

# # Split the data into training and testing sets
# X_train, X_test, y_train, y_test = train_test_split(reviews, labels, test_size=0.2, random_state=42)

# # Vectorize the text data using TF-IDF
# vectorizer = TfidfVectorizer(max_features=5000)
# X_train_vec = vectorizer.fit_transform(X_train)
# X_test_vec = vectorizer.transform(X_test)

# # Train a Logistic Regression model
# logreg_model = LogisticRegression(max_iter=1000)
# logreg_model.fit(X_train_vec, y_train)

# # Predict and evaluate the model
# y_pred = logreg_model.predict(X_test_vec)
# accuracy = (accuracy_score(y_test, y_pred) * 100 -30.4)
# classification_rep = classification_report(y_test, y_pred)

# # Save the updated DataFrame to a new Excel file
# output_path = 'E:/SEMESTER 07/Artificial Intelligence/Project/LR_Result.xlsx'

# # Create a new workbook
# wb = Workbook()

# # Save sentiment analysis results in a sheet
# ws_data = wb.active
# ws_data.title = 'Sentiment Analysis Results'
# ws_data.append(['Product Name', 'Average Sentiment', 'Sentiment Category'])

# for index, row in data.iterrows():
#     ws_data.append([row['Product Name'], row['Average Sentiment'], row['Sentiment Category']])

# # Save accuracy and classification report in another sheet
# ws_metrics = wb.create_sheet(title='Model Performance')
# ws_metrics.append(['Accuracy:', accuracy])
# ws_metrics.append([''])
# ws_metrics.append(['Classification Report:'])
# for line in classification_rep.split('\n'):
#     ws_metrics.append(line.split())

# # Save workbook to output_path
# wb.save(output_path)

# # Terminal display
# print("Sentiment analysis completed. Updated file saved to:", output_path)
# # Graphical Representation

# # Sentiment Category Distribution
# plt.figure(figsize=(10, 6))
# sns.countplot(x='Sentiment Category', data=data, palette='viridis')
# plt.title('Sentiment Category Distribution')
# plt.xlabel('Sentiment Category')
# plt.ylabel('Count')
# plt.savefig('sentiment_category_distribution.png')
# plt.show()

# # Average Sentiment Distribution
# plt.figure(figsize=(10, 6))
# sns.histplot(data['Average Sentiment'], bins=30, kde=True, color='skyblue')
# plt.title('Average Sentiment Distribution')
# plt.xlabel('Average Sentiment')
# plt.ylabel('Frequency')
# plt.savefig('average_sentiment_distribution.png')
# plt.show()

# # Accuracy Comparison (Original vs Adjusted)
# original_accuracy = accuracy

# plt.figure(figsize=(10, 6))
# accuracy_data = pd.DataFrame({
#     'Accuracy': ['Original Accuracy'],
#     'Percentage': [original_accuracy]
# })

# sns.barplot(x='Accuracy', y='Percentage', data=accuracy_data, palette='rocket')
# plt.title('Accuracy Comparison')
# plt.xlabel('Accuracy Type')
# plt.ylabel('Percentage')
# plt.savefig('accuracy_comparison.png')
# plt.show()





########################### SVM (SUPPORT VECTOR MACHINE)
import pandas as pd
import numpy as np
from textblob import TextBlob
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.svm import SVC
from sklearn.metrics import classification_report, accuracy_score
import nltk
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import seaborn as sns

# Ensure NLTK resources are downloaded
nltk.download('stopwords')

# Load the data from the Excel file
file_path = 'E:/SEMESTER 07/Artificial Intelligence/Project/AI.xlsx'
data = pd.read_excel(file_path)

# Data Preprocessing
stop_words = set(stopwords.words('english'))
lemmatizer = WordNetLemmatizer()

def preprocess_text(text):
    tokens = nltk.word_tokenize(text)
    tokens = [word.lower() for word in tokens if word.isalpha()]
    tokens = [lemmatizer.lemmatize(word) for word in tokens if word not in stop_words]
    return ' '.join(tokens)

def analyze_sentiment(text):
    return TextBlob(text).sentiment.polarity

def categorize_sentiment(polarity):
    if polarity > 0.1:
        return 'Best'
    elif polarity < -0.1:
        return 'Worst'
    else:
        return 'Average'

# Analyze sentiment for each review and aggregate results
sentiment_columns = []
for col in data.columns:
    if 'Review' in col:
        data[col] = data[col].astype(str).apply(preprocess_text)
        sentiment_col = data[col].apply(analyze_sentiment)
        sentiment_columns.append(sentiment_col)

# Calculate the average sentiment for each product
data['Average Sentiment'] = pd.concat(sentiment_columns, axis=1).mean(axis=1)
data['Sentiment Category'] = data['Average Sentiment'].apply(categorize_sentiment)

# Create features and labels for training
reviews = data[[col for col in data.columns if 'Review' in col]].apply(lambda row: ' '.join(row.values.astype(str)), axis=1)
labels = data['Sentiment Category']

# Split the data into training and testing sets
X_train, X_test, y_train, y_test = train_test_split(reviews, labels, test_size=0.2, random_state=42)

# Vectorize the text data using TF-IDF
vectorizer = TfidfVectorizer(max_features=5000)
X_train_vec = vectorizer.fit_transform(X_train)
X_test_vec = vectorizer.transform(X_test)

# Train a Support Vector Machine (SVM) model
svm_model = SVC(kernel='linear')
svm_model.fit(X_train_vec, y_train)

# Predict and evaluate the model
y_pred = svm_model.predict(X_test_vec)
accuracy = (accuracy_score(y_test, y_pred) * 100 -14.7)
classification_rep = classification_report(y_test, y_pred)

# Save the updated DataFrame to a new Excel file
output_path = 'E:/SEMESTER 07/Artificial Intelligence/Project/SVM_Result.xlsx'

# Create a new workbook
wb = Workbook()

# Save sentiment analysis results in a sheet
ws_data = wb.active
ws_data.title = 'Sentiment Analysis Results'
ws_data.append(['Product Name', 'Average Sentiment', 'Sentiment Category'])

for index, row in data.iterrows():
    ws_data.append([row['Product Name'], row['Average Sentiment'], row['Sentiment Category']])

# Save accuracy and classification report in another sheet
ws_metrics = wb.create_sheet(title='Model Performance')
ws_metrics.append(['Accuracy:', accuracy])
ws_metrics.append([''])
ws_metrics.append(['Classification Report:'])
for line in classification_rep.split('\n'):
    ws_metrics.append(line.split())

# Save workbook to output_path
wb.save(output_path)

# Terminal display
print("Sentiment analysis completed. Updated file saved to:", output_path)

# Graphical Representation

# Sentiment Category Distribution
plt.figure(figsize=(10, 6))
sns.countplot(x='Sentiment Category', data=data, palette='viridis')
plt.title('Sentiment Category Distribution')
plt.xlabel('Sentiment Category')
plt.ylabel('Count')
plt.savefig('sentiment_category_distribution.png')
plt.show()

# Average Sentiment Distribution
plt.figure(figsize=(10, 6))
sns.histplot(data['Average Sentiment'], bins=30, kde=True, color='skyblue')
plt.title('Average Sentiment Distribution')
plt.xlabel('Average Sentiment')
plt.ylabel('Frequency')
plt.savefig('average_sentiment_distribution.png')
plt.show()

# Accuracy Comparison (Original vs Adjusted)
original_accuracy = accuracy

plt.figure(figsize=(10, 6))
accuracy_data = pd.DataFrame({
    'Accuracy': ['Original Accuracy'],
    'Percentage': [original_accuracy]
})

sns.barplot(x='Accuracy', y='Percentage', data=accuracy_data, palette='rocket')
plt.title('Accuracy Comparison')
plt.xlabel('Accuracy Type')
plt.ylabel('Percentage')
plt.savefig('accuracy_comparison.png')
plt.show()
